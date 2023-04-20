import openpyxl

# file = "G:\INFOSYS Lectures & Codes\Python Testing Codes\Python Codes\Empty.xlsx"
file = "G:\INFOSYS Lectures & Codes\Python Testing Codes\Python Codes\MutlipleData.xlsx"
workBook = openpyxl.load_workbook(file)  # It will load the workbook
sheet = workBook.active  # Get active sheet from Excel if we have only one sheet then we can write "ACTIVE" , else if we have no of sheets then
                         # we can specify as sheet = workbook["Sheet1"]


# This loop can write the same data in all rows and cols
# for r in range(1, 6):
#     for c in range(1, 4):
#         sheet.cell(r, c).value = "Welcome"


# To write multiple data

sheet.cell(1,1).value = "EMP_ID"
sheet.cell(1,2).value = "EMP_Name"
sheet.cell(1,3).value = "Salary"


sheet.cell(2,1).value = 15512
sheet.cell(2,2).value = "Rohan Das"
sheet.cell(2,3).value = 78000

sheet.cell(3,1).value = 78812
sheet.cell(3,2).value = "Mohit Saxena"
sheet.cell(3,3).value = 45000

sheet.cell(4,1).value = 44562
sheet.cell(4,2).value = "Anjali Sangwan"
sheet.cell(4,3).value = 150000


workBook.save(file)
