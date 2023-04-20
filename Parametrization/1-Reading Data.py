import openpyxl

file = "G:\INFOSYS Lectures & Codes\Python Testing Codes\Python Codes\Data.xlsx"
workBook = openpyxl.load_workbook(file)  # It will load the workbook
sheet = workBook["Sheet1"]

rows = sheet.max_row  # Count no of rows in Excel sheets
cols = sheet.max_column  # Count no of columns in Excel sheets

# Reading all the rows and columns from Excel sheet
for r in range(1,rows+1):
    for c in range(1, cols+1):
        print(sheet.cell(r,c).value, end = "                             ")
    print()
